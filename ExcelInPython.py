from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "가족정보" #파일 이름 지정
ws.merge_cells('A1:E1') # 셀 합치기 ** 절대 띄어쓰기 하지 말것 ['A1 : E1'] 안됨
ws['A1'] = "가족정보" # 셀 A1에 가족정보 넣기
column = ['이름','전호번호','소속'] # 다음셀에 넣을 리스트 작성 하나의 셀에 하나씩 지정됨
ws.append(column)  # 리스트를 포함
row = ['유인목','010-2007-9719', '가족'] # 다음칸에 넣을 리스트  위에 맞춰서 하나씩 작성
ws.append(row)# 리스트에 포함
wb.create_sheet('오빠네가족') #새로운 시트 작성
wb.create_sheet('장인어른과 장모님') #새로운 시트 작성
wb.save("D:\파이썬 수업\git\pythonstart\가족정보.xlsx") # 엑셀이 저정될 위치 지정
wb.close() # 마무리
